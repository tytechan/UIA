from openpyxl.styles import Border, Side, Font, Alignment
from config.ErrConfig import CNBMException, handleErr
from collections import Iterable
import openpyxl
import win32com.client
import re

class ParseExcel:
    def __init__(self):
        self.workbook = None
        self.excelFile = None

    @CNBMException
    def loadWorkBook(self, filaPath):
        ''' 将excel文件加载到内存，并获取其workbook对象
        :param filaPath: 文件路径
        '''
        try:
            self.workbook = openpyxl.load_workbook(filaPath)
            self.excelFile = filaPath
            return ParseSheet(self.workbook, self.excelFile)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def create(self, filePath):
        ''' 创建新工作簿并保存
        :param filePath: 文件保存路径
        '''
        try:
            self.workbook = openpyxl.Workbook()
            self.excelFile = filePath
            self.workbook.save(self.excelFile)
            return ParseSheet(self.workbook, self.excelFile)
        except Exception as e:
            handleErr(e)
            raise e


class ParseSheet:
    def __init__(self, workbook, excelFile):
        self.workbook = workbook
        self.excelFile = excelFile
        self.sheet = None

    @CNBMException
    def createSheet(self, title=None, index=None):
        ''' 新建sheet
        :param title: sheet名，选填
        :param index: sheet序号
        :return:
        '''
        try:
            self.sheet = self.workbook.create_sheet(title=title, index=index)
            # self.save()
            return ParseCell(self.workbook, self.excelFile, self.sheet)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def getSheet(self, sheetName=None, sheetIndex=0):
        ''' 获取sheet对象（若名称不为空，则优先通过名称查找；否则通过索引号查找，默认获取第一个sheet）
        :param sheetName: sheet名称
        :param sheetIndex: sheet序号
        '''
        try:
            if sheetName:
                self.sheet = self.workbook[sheetName]
                # sheet = self.workbook.get_sheet_by_name(sheetName)
            else:
                self.sheet = self.workbook.worksheets[sheetIndex]
            return ParseCell(self.workbook, self.excelFile, self.sheet)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def remove(self, sheetName=None, sheetIndex=None):
        ''' 删除选定sheet
        （若名称不为空，则优先通过名称查找；否则通过索引号查找，默认获取第一个sheet）
        :param sheetName: sheet名称
        :param sheetIndex: sheet序号
        '''
        try:
            self.getSheet(sheetName, sheetIndex)
            self.workbook.remove(self.sheet)
            # self.save()
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def rename(self, newName, sheetName=None, sheetIndex=None):
        ''' 重命名选定sheet
        （若名称不为空，则优先通过名称查找；否则通过索引号查找，默认获取第一个sheet）
        :param newName: 新定义的sheet名称
        :param sheetName: sheet名称
        :param sheetIndex: sheet序号
        '''
        try:
            self.getSheet(sheetName, sheetIndex)
            self.sheet.title = newName
            # self.save()
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def runVBA(self, func, visible=True):
        ''' 在 self.workbook 中运行VBA
        :param func: 函数名
        :param visible: 运行过程是否可见，默认可见
        '''
        try:
            xlApp = win32com.client.DispatchEx("Excel.Application")
            xlApp.Visible = visible
            xlApp.DisplayAlerts = 0
            xlBook = xlApp.Workbooks.Open(self.excelFile, False)
            xlBook.Application.Run(func)
            xlBook.Close(True)
            xlApp.quit()
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def sheets(self):
        '''
        返回 self.workbook 中所有sheet列表（list）
        '''
        try:
            return self.workbook.worksheets
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def save(self, savePath=None):
        '''
        保存文件
        '''
        try:
            path = savePath if savePath else self.excelFile
            self.workbook.save(path)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def close(self, save=True, savePath=None):
        '''
        关闭文件，默认保存
        '''
        try:
            if save:
                self.save(savePath)
            self.workbook.close()
        except Exception as e:
            handleErr(e)
            raise e


class ParseCell:
    def __init__(self, workbook, excelFile, sheet):
        self.workbook = workbook
        self.excelFile = excelFile
        self.sheet = sheet
        #颜色对应的RGB值
        self.RGBDict = {'red':'FFFF3030', 'green':'FF008B00'}

    def getValue(self, rng):
        '''
        获取区域中所有单元格的值，内部函数
        :param rng: tuple类型
        :return: 单元格value的多维数组
        '''
        try:
            cellsValue = list()
            for cells in rng:
                row = list()
                if isinstance(cells, Iterable):
                    # 可迭代
                    for cell in cells:
                        row.append(cell.value)
                else:
                    row = cells.value
                cellsValue.append(row)
            return cellsValue
        except Exception as e:
            raise e

    def delNone(self, cells):
        '''
        删除二维value数组首位的空值，内部函数
        :param cells: list类型
        :return: 单元格value的二维数组
        '''
        try:
            for i in range(len(cells)-1, -1, -1):
                if cells[i]:
                    break
            for j in range(len(cells)):
                if cells[j]:
                    break
            myList = []
            while j <= i:
                myList.append(cells[j])
                j += 1
            return myList
        except Exception as e:
            raise e

    @CNBMException
    def rowCount(self):
        '''
        获取sheet中有数据区域的结束行号
        '''
        try:
            return self.sheet.max_row
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def colCount(self):
        '''
        获取sheet中有数据区域的结束列号
        '''
        try:
            return self.sheet.max_column
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def minRowNum(self):
        '''
        获取sheet中有数据区域的开始行号
        '''
        try:
            return self.sheet.min_row
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def minColNum(self):
        '''
        获取sheet中有数据区域的开始列号
        '''
        try:
            return self.sheet.min_column
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def row(self, rowNo, getValue=True, keepNone=False):
        ''' 获取sheet中某一行
        :param rowNo: 行号（从1开始）
        :param getValue: True则返回单元格输值，False则返回tuple数据
        :param keepNone: 返回数值时，False则删除首尾空值，True则保留
        '''
        try:
            myrows = self.sheet[rowNo]
            if getValue:
                myrows = self.getValue(myrows)
                if not keepNone:
                    myrows = self.delNone(myrows)
            return myrows

            # colNo = ParseExcel().getColsNumber(sheet)
            # print("colNo:",colNo)
            # MyValue = []
            # for i in range(colNo):
            #     MyValue = MyValue.append(sheet.cell(rowNo,i).value)
            # return MyValue
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def column(self, colNo, getValue=True, keepNone=False):
        ''' 获取sheet中某一列，返回该列所有数据内容组成的tuple
        :param colNo: 列号，可填列标或列序号（从1开始）
        如：sheet.column(2) 或 sheet.column("B")
        '''
        try:
            if isinstance(colNo, int):
                maxRow = self.sheet.max_row
                cells = list()
                for i in range(maxRow):
                    cell = self.sheet.cell(row=i+1, column=colNo)
                    cells.append(cell)
                    # if cell.value:
                    #     cells.append(cell)
                cells = tuple(cells)
            else:
                cells = self.sheet[colNo]

            if getValue:
                cells = self.getValue(cells)
                if not keepNone:
                    cells = self.delNone(cells)
            return cells
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def delete_rows(self, idx, amount=1):
        '''
        删除行
        :param idx: 起始行号
        :param amount: 从idx开始数共删除的行数
        '''
        try:
            self.sheet.delete_rows(idx, amount)
            # self.workbook.save(self.excelFile)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def delete_cols(self, idx, amount=1):
        '''
        删除列
        :param idx: 起始列号（数值型）
        :param amount: 从idx开始数共删除的列数
        '''
        try:
            self.sheet.delete_cols(idx, amount)
            # self.workbook.save(self.excelFile)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def read(self, coordinate=None, rowNo=None, colsNo=None):
        '''
        根据单元格所在的位置索引获取该单元格中的值
        sheet.cell(row = 1,column = 1).value 表示excel中第一行第一列的值
        :param coordinate: 单元格坐标（“A3”、"B5"）
        :param rowNo: 行号（从1开始）
        :param colsNo: 列号（从1开始）
        '''
        try:
            if coordinate != None:
                # coordinate = coordinate.decode('utf-8')
                return self.sheet[coordinate].value
            elif coordinate is None and rowNo is not None and colsNo is not None:
                return self.sheet.cell(row=rowNo, column=colsNo).value
        except Exception as e:
            handleErr(e)
            raise e


    @CNBMException
    def cell(self, coordinate=None, rowNo=None, colsNo=None):
        '''
        获取某个单元格对象，可以根据单元格所在位置的数字索引
        也可以直接根据excel中单元格的编码及坐标，
        如：getCell(sheet,coordinate = 'A1')
        或：getCell(sheet,rowNo = 1,colsNo = 2)
        :param coordinate: 单元格坐标（“A3”、"B5"）
        :param rowNo: 行号（从1开始）
        :param colsNo: 列号（从1开始）
        '''
        try:
            if coordinate != None:
                # coordinate = coordinate.decode('utf-8')
                return self.sheet[coordinate]
                # return sheet.cell(coordinate = coordinate)
            elif coordinate == None and rowNo is not None and colsNo is not None:
                return self.sheet.cell(row=rowNo,column=colsNo)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def range(self, startRng=None, endRng=None, getValue=True):
        '''
        获取多个单元格对象，返回tuple类型数据
        range("B2", "C3")，获取“B2”到“C3区域”
        range("B", "C")，获取“B”列到“C”列
        range(2, 3)，获取2-3行
        :param startRng: 起始区域
        :param endRng: 终止区域
        '''
        try:
            if not startRng and not endRng:
                startRng = "A1"
                rowCount = self.rowCount()
                colCount = self.colCount()
                # chr(97)='a'
                endRng = "%s%d" %(chr(96+colCount), rowCount)

            if isinstance(startRng, str) and isinstance(endRng, str):
                runStr = 'self.sheet["%s":"%s"]' %(startRng, endRng)
            else:
                runStr = 'self.sheet[%s:%s]' %(startRng, endRng)
            cells = eval(runStr)

            if getValue:
                cells = self.getValue(cells)
            return cells
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def write(self, content, coordinate=None, rowNo=None, colsNo=None, style=None):
        '''
        根据单元格在excel中的编码坐标或者数字索引坐标向单元格中写入数据（写之前确认文件关闭），
        下表从1开始，参数style表示字体的颜色名称，如：red、green
        :param content: 输入内容
        :param coordinate: 单元格坐标（“A3”、"B5"）
        :param rowNo: 行号（从1开始）
        :param colsNo: 列号（从1开始）
        :param style: 字体的颜色名称（“red”、“green”，默认为黑色）
        '''
        try:
            if coordinate is not None:
                # coordinate = coordinate.decode('utf-8')
                self.sheet[coordinate].value = content
                # sheet.cell(coordinate = coordinate).value = content
                if style is not None:
                    self.sheet[coordinate].font = Font(color = self.RGBDict[style])
                    # sheet.cell(coordinate = coordinate).font = Font(color = self.RGBDict[style])
            elif coordinate == None and rowNo is not None and colsNo is not None:
                # sheet.cell(row = rowNo,column = colsNo).value = ""
                self.sheet.cell(row=rowNo, column=colsNo).value = content
                if style:
                    self.sheet.cell(row=rowNo, column=colsNo).font = Font(color=self.RGBDict[style])
            # self.workbook.save(self.excelFile)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def alignment(self, startRng, endRng=None, horizontal="center", vertical="center"):
        '''
        设置单个或多个单元格的对齐方式
        处理“.xlsm”文件可能导致文件损坏，慎用；处理“.xlsx”正常
        sht.alignment("B2", horizontal="right", vertical="top")
        sht.alignment("B2", "C3")
        sht.alignment(2, 3, horizontal="left", vertical="bottom")
        sht.alignment("B", "C", horizontal="right", vertical="top")
        :param startRng: 起始区域
        :param endRng: 终止区域（选填）
        :param horizontal: 水平对齐方式（"general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"）
        :param vertical: 垂直对齐方式（"top", "center", "bottom", "justify", "distributed"）
        '''
        try:
            if endRng:
                if isinstance(startRng, str) and isinstance(endRng, str):
                    runStr = 'self.sheet["%s":"%s"]' %(startRng, endRng)
                else:
                    runStr = 'self.sheet[%s:%s]' %(startRng, endRng)
                rng = eval(runStr)
                for i in range(len(rng)):
                    for cell in rng[i]:
                        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
            else:
                self.sheet[startRng].alignment = Alignment(horizontal=horizontal, vertical=vertical)
            # self.workbook.save(self.excelFile)
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def height(self, rowNum, height):
        '''
        设置单行的单元格高度
        :param rowNum: 行序号（从1开始）
        :param height: 高度
        '''
        try:
            self.sheet.row_dimensions[rowNum].height = height
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def width(self, colNum, width):
        '''
        设置单列的单元格宽度
        :param colNum: 列序号
        :param width: 宽度
        '''
        try:
            self.sheet.column_dimensions[colNum].width = width
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def merge(self, startRng, endRng):
        '''
        合并单元格
        :param startRng: 起始区域
        :param endRng: 终止区域
        '''
        try:
            self.sheet.merge_cells('%s:%s' %(startRng, endRng))
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def unmerge(self, startRng, endRng):
        '''
        拆分单元格
        :param startRng: 起始区域
        :param endRng: 终止区域
        '''
        try:
            self.sheet.unmerge_cells('%s:%s' %(startRng, endRng))
        except Exception as e:
            handleErr(e)
            raise e

    @CNBMException
    def paste(self, data, startRng=None):
        '''
        复制单元格区域
        :param data: （tuple）原始数据单元格
        :param startRng: 新表中粘贴的起始单元格（不填则与原表位置一致进行粘贴）
        '''
        try:
            initRng = data[0][0]
            # 计算偏移量
            if not startRng:
                rowMove = 0
                colMove = 0
            else:
                rowNum = int(re.findall('\d+', startRng)[0])
                # ord("A") = 65
                colNum = ord(startRng.replace(str(rowNum), "")) - 64
                rowMove = rowNum - initRng.row
                colMove = colNum - initRng.column

            for i in range(len(data)):
                for cell in data[i]:
                    row = cell.row + rowMove
                    column = cell.column + colMove
                    self.sheet.cell(row=row, column=column).value = cell.value
        except Exception as e:
            handleErr(e)
            raise e


if __name__ == "__main__":
    # 调试前先注释 @CNBMException
    Excel = ParseExcel()
    # 创建新工作簿
    # wb0 = Excel.create(r"E:\python相关\RPA\UIA\localSDK\abc.xlsx")

    # 操作已有工作簿
    wb = Excel.loadWorkBook(r"E:\python相关\RPA\UIA\localSDK\test.xlsx")
    print("所有sheet列表：", wb.sheets())

    sht = wb.getSheet("Sheet1")
    print("最大行数：%s" %sht.rowCount())
    print("最小行数：%s" %sht.minRowNum())
    print("最大列数：%s" %sht.colCount())
    print("最小列数：%s\n" %sht.minColNum())

    # 删除行列
    sht.delete_rows(9)
    sht.delete_cols(8, 2)

    cell = sht.cell("B3")
    print("B3值：", cell.value)
    cell1 = sht.cell(rowNo=3, colsNo=3)
    print("C3值：", cell1.value)
    print("C2值：", sht.read("C2"))
    print("D2的值：", sht.read(rowNo=2, colsNo=4), "\n")

    col2 = sht.column("B", getValue=False)
    print("第二列为：", col2)
    row3 = sht.row(3, getValue=False)
    print("第三行为：", row3)
    print("C3的值：", row3[2].value, "\n")

    col2 = sht.column(2)
    print("第二列有值项为：", col2)
    row3 = sht.row(3)
    print("第三行有值项为：", row3)
    row3 = sht.row(3, keepNone=True)
    print("第三行所有值为：", row3, "\n")

    print("B2:C3为：", sht.range("B2", "C3", getValue=False))
    print("B2:C3值为：", sht.range("B2", "C3"))
    print("B:C为：", sht.range("B", "C", getValue=False))
    print("B:C值为：", sht.range("B", "C"))
    print("2-3行为：", sht.range(2, 3, getValue=False))
    print("2-3行值为：", sht.range(2, 3))
    print("获取整个sheet为：", sht.range(getValue=False))
    print("获取整个sheet值为：", sht.range())

    # 设置属性
    sht.alignment("B2", horizontal="right", vertical="top")
    sht.alignment("B2", "C3")
    sht.alignment(2, 3, horizontal="left", vertical="bottom")
    sht.alignment("B", "C", horizontal="right", vertical="top")

    sht.height(6, 50)
    sht.width("G", 100)

    # 拆分合并
    sht.merge("D3", "E4")
    sht.unmerge("D3", "E4")

    # 写值
    sht.write("test", "G4")
    sht.write("test1", rowNo=7, colsNo=5)

    # 创建sheet
    sht1 = wb.createSheet("test")
    sht2 = wb.createSheet("test1", 0)

    # 复制
    data = sht.range("B2", "D3")
    sht1.paste(data)
    sht2.paste(data, "D5")

    # # 删除sheet
    # wb.remove("test")
    # wb.remove(sheetIndex=0)
    #
    # # 重命名sheet
    # wb.rename("重命名", "Sheet1")
    # wb.rename("Sheet1", sheetIndex=0)

    # 运行vba
    # wb.runVBA("test")

    # wb.save()
    wb.close()